"""Excel file handler with formatting support for ReportBot."""
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from src.models import Business, Review


# Column configuration
COLUMNS = {
    'url': {'header': 'URL', 'width': 45},
    'business_name': {'header': 'İşletme Adı', 'width': 25},
    'count': {'header': 'Adet', 'width': 8},
    'report_id': {'header': 'Rapor ID', 'width': 20},
    'review_url': {'header': 'Yorum URL', 'width': 45},
    'report_date': {'header': 'Rapor Tarihi', 'width': 12},
    'reviewer_name': {'header': 'Yorumcu', 'width': 20},
    'review_text': {'header': 'Yorum Metni', 'width': 50},
    'rating': {'header': 'Puan', 'width': 8},
    'status': {'header': 'Durum', 'width': 15},
}

COLUMN_ORDER = ['url', 'count', 'business_name', 'report_id', 'review_url', 'report_date', 'reviewer_name', 'review_text', 'rating', 'status']

# Columns to merge for multiple reviews of the same restaurant
MERGE_COLUMNS = ['url', 'count', 'report_id']

# Status options for dropdown
STATUS_OPTIONS = ['beklemede', 'silindi', 'reddedildi', 'işleniyor']

# Styling
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')  # Dark blue
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

ALT_ROW_FILL = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')  # Light blue
NORMAL_ALIGNMENT = Alignment(vertical='top', wrap_text=True)
MERGED_CELL_ALIGNMENT = Alignment(vertical='center', horizontal='left', wrap_text=True)  # For merged cells

THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)

# Status colors
STATUS_COLORS = {
    'beklemede': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),  # Yellow
    'silindi': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),  # Green
    'reddedildi': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),  # Red
    'işleniyor': PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid'),  # Blue
}


def create_formatted_workbook() -> Workbook:
    """Create a new workbook with formatted headers and validation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "URLs"
    
    # Write headers
    for col_idx, col_key in enumerate(COLUMN_ORDER, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = COLUMNS[col_key]['header']
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMNS[col_key]['width']
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add status dropdown validation (for rows 2 to 1000)
    status_col_idx = COLUMN_ORDER.index('status') + 1
    status_validation = DataValidation(
        type='list',
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=True
    )
    status_validation.error = 'Lütfen listeden bir değer seçin'
    status_validation.errorTitle = 'Geçersiz Değer'
    status_validation.prompt = 'Durum seçin'
    status_validation.promptTitle = 'Durum'
    
    status_col_letter = get_column_letter(status_col_idx)
    status_validation.add(f'{status_col_letter}2:{status_col_letter}1000')
    ws.add_data_validation(status_validation)
    
    # Set row height for header
    ws.row_dimensions[1].height = 25
    
    return wb


def apply_row_formatting(ws, row_idx: int, is_alt_row: bool = False):
    """Apply formatting to a data row."""
    for col_idx in range(1, len(COLUMN_ORDER) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.alignment = NORMAL_ALIGNMENT
        cell.border = THIN_BORDER
        
        if is_alt_row:
            cell.fill = ALT_ROW_FILL
    
    # Apply status-specific coloring
    status_col_idx = COLUMN_ORDER.index('status') + 1
    status_cell = ws.cell(row=row_idx, column=status_col_idx)
    status_value = str(status_cell.value or '').lower().strip()
    
    if status_value in STATUS_COLORS:
        status_cell.fill = STATUS_COLORS[status_value]


def merge_cells_for_group(ws, start_row: int, end_row: int, col_indices: dict):
    """Merge cells for common columns when there are multiple reviews.
    
    Args:
        ws: Worksheet object.
        start_row: First row of the group.
        end_row: Last row of the group.
        col_indices: Dictionary mapping column keys to column indices.
    """
    if start_row >= end_row:
        return  # Nothing to merge
    
    for col_key in MERGE_COLUMNS:
        if col_key in col_indices:
            col_idx = col_indices[col_key]
            col_letter = get_column_letter(col_idx)
            
            # Merge cells
            merge_range = f'{col_letter}{start_row}:{col_letter}{end_row}'
            ws.merge_cells(merge_range)
            
            # Apply centered alignment to the merged cell
            cell = ws.cell(row=start_row, column=col_idx)
            cell.alignment = MERGED_CELL_ALIGNMENT
            cell.border = THIN_BORDER


def read_excel_urls_with_count(excel_path: str) -> list[tuple[str, int]]:
    """Read URLs and review counts from an Excel file.
    
    Only returns rows where 'report_id' is empty (not yet processed).
    
    Args:
        excel_path: Path to the Excel file.
        
    Returns:
        List of (url, count) tuples.
        
    Raises:
        FileNotFoundError: If the Excel file doesn't exist.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    # Find column indices from header row
    col_indices = {}
    for col_idx, cell in enumerate(ws[1], 1):
        header_value = str(cell.value or '').lower().strip()
        for col_key, col_config in COLUMNS.items():
            if header_value == col_config['header'].lower() or header_value == col_key:
                col_indices[col_key] = col_idx
                break
    
    if 'url' not in col_indices:
        wb.close()
        return []
    
    data = []
    for row_idx in range(2, ws.max_row + 1):
        url_cell = ws.cell(row=row_idx, column=col_indices['url'])
        url = str(url_cell.value or '').strip()
        
        if not url:
            continue
        
        # Only skip rows that have report_id (this specific row is processed)
        # Same URL can be processed multiple times in different rows
        if 'report_id' in col_indices:
            report_id_cell = ws.cell(row=row_idx, column=col_indices['report_id'])
            if report_id_cell.value:
                continue
        
        # Get count, default to 1
        count = 1
        if 'count' in col_indices:
            count_cell = ws.cell(row=row_idx, column=col_indices['count'])
            if count_cell.value:
                try:
                    count = int(count_cell.value)
                except (ValueError, TypeError):
                    count = 1
        
        data.append((url, count))
    
    wb.close()
    return data


def update_excel_with_report(
    excel_path: str,
    url: str,
    report_id: str,
    reviews: list[Review] = None,
    business: Business = None
) -> bool:
    """Update the Excel file with report ID and review details.
    
    Args:
        excel_path: Path to the Excel file.
        url: The URL to find and update.
        report_id: The report ID to save.
        reviews: List of Review objects that were reported.
        business: Business object for the place.
        
    Returns:
        True if update was successful, False otherwise.
    """
    if not os.path.exists(excel_path):
        return False
    
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Find column indices
        col_indices = {}
        for col_idx, cell in enumerate(ws[1], 1):
            header_value = str(cell.value or '').lower().strip()
            for col_key, col_config in COLUMNS.items():
                if header_value == col_config['header'].lower() or header_value == col_key:
                    col_indices[col_key] = col_idx
                    break
        
        if 'url' not in col_indices:
            wb.close()
            return False
        
        # Find the row with matching URL that hasn't been processed yet
        url_row_idx = None
        for row_idx in range(2, ws.max_row + 1):
            url_cell = ws.cell(row=row_idx, column=col_indices['url'])
            if str(url_cell.value or '').strip() == url:
                # Check if this row already has a report_id (skip if processed)
                if 'report_id' in col_indices:
                    report_id_cell = ws.cell(row=row_idx, column=col_indices['report_id'])
                    if report_id_cell.value:
                        continue  # Skip already processed rows
                url_row_idx = row_idx
                break
        
        if url_row_idx is None:
            wb.close()
            return False
        
        # Update the row with report data
        if 'report_id' in col_indices:
            ws.cell(row=url_row_idx, column=col_indices['report_id']).value = report_id
        
        if business and 'business_name' in col_indices:
            ws.cell(row=url_row_idx, column=col_indices['business_name']).value = business.name or ''
        
        # Add first review details to this row
        today_str = datetime.now().strftime('%d.%m.%Y')
        
        if reviews and len(reviews) > 0:
            first_review = reviews[0]
            if 'review_url' in col_indices:
                ws.cell(row=url_row_idx, column=col_indices['review_url']).value = first_review.review_url or ''
            if 'report_date' in col_indices:
                ws.cell(row=url_row_idx, column=col_indices['report_date']).value = today_str
            if 'reviewer_name' in col_indices:
                ws.cell(row=url_row_idx, column=col_indices['reviewer_name']).value = first_review.author_name or ''
            if 'review_text' in col_indices:
                ws.cell(row=url_row_idx, column=col_indices['review_text']).value = first_review.text or ''
            if 'rating' in col_indices:
                ws.cell(row=url_row_idx, column=col_indices['rating']).value = first_review.rating
            if 'status' in col_indices:
                ws.cell(row=url_row_idx, column=col_indices['status']).value = 'beklemede'
        
        # Insert additional rows for remaining reviews FIRST (before formatting)
        num_reviews = len(reviews) if reviews else 1
        
        if reviews and len(reviews) > 1:
            for i, review in enumerate(reviews[1:], 1):
                new_row_idx = url_row_idx + i
                ws.insert_rows(new_row_idx)
                
                if business and 'business_name' in col_indices:
                    ws.cell(row=new_row_idx, column=col_indices['business_name']).value = business.name or ''
                if 'review_url' in col_indices:
                    ws.cell(row=new_row_idx, column=col_indices['review_url']).value = review.review_url or ''
                if 'report_date' in col_indices:
                    ws.cell(row=new_row_idx, column=col_indices['report_date']).value = today_str
                if 'reviewer_name' in col_indices:
                    ws.cell(row=new_row_idx, column=col_indices['reviewer_name']).value = review.author_name or ''
                if 'review_text' in col_indices:
                    ws.cell(row=new_row_idx, column=col_indices['review_text']).value = review.text or ''
                if 'rating' in col_indices:
                    ws.cell(row=new_row_idx, column=col_indices['rating']).value = review.rating
                if 'status' in col_indices:
                    ws.cell(row=new_row_idx, column=col_indices['status']).value = 'beklemede'
        
        # Apply formatting to all rows in the group
        for i in range(num_reviews):
            row_idx = url_row_idx + i
            is_alt_row = (row_idx % 2 == 0)
            apply_row_formatting(ws, row_idx, is_alt_row)
        
        # Merge common cells if there are multiple reviews
        if num_reviews > 1:
            end_row = url_row_idx + num_reviews - 1
            merge_cells_for_group(ws, url_row_idx, end_row, col_indices)
        
        wb.save(excel_path)
        wb.close()
        
        print(f"✅ Excel güncellendi: {url[:50]}... -> {report_id}")
        if reviews:
            print(f"   Raporlanan yorumlar ({len(reviews)}) kaydedildi.")
        return True
        
    except Exception as e:
        print(f"⚠️ Excel güncellenirken hata: {e}")
        return False


def convert_csv_to_excel(csv_path: str, excel_path: str = None) -> str:
    """Convert a CSV file to a formatted Excel file.
    
    Args:
        csv_path: Path to the source CSV file.
        excel_path: Optional path for the output Excel file.
                   If not provided, replaces .csv extension with .xlsx
                   
    Returns:
        Path to the created Excel file.
    """
    import csv
    
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    
    if excel_path is None:
        excel_path = str(Path(csv_path).with_suffix('.xlsx'))
    
    # Read CSV data
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        rows = list(reader)
    
    if not rows:
        raise ValueError("CSV file is empty")
    
    # Create formatted workbook
    wb = create_formatted_workbook()
    ws = wb.active
    
    # Detect CSV columns
    csv_header = [col.lower().strip() for col in rows[0]]
    has_header = 'url' in csv_header
    
    # Map CSV columns to our columns
    csv_to_col = {}
    if has_header:
        for csv_idx, csv_col in enumerate(csv_header):
            for col_key in COLUMN_ORDER:
                if csv_col == col_key or csv_col == COLUMNS[col_key]['header'].lower():
                    csv_to_col[csv_idx] = COLUMN_ORDER.index(col_key) + 1
                    break
        data_rows = rows[1:]
    else:
        # Assume first column is URL
        csv_to_col[0] = 1
        data_rows = rows
    
    # Build column indices mapping
    col_indices = {col_key: COLUMN_ORDER.index(col_key) + 1 for col_key in COLUMN_ORDER}
    
    # Write data rows and track groups for merging
    groups = []  # List of (start_row, end_row) tuples
    current_group_start = None
    url_col_in_csv = None
    
    # Find URL column index in CSV
    for csv_idx, excel_col_idx in csv_to_col.items():
        if excel_col_idx == 1:  # URL column
            url_col_in_csv = csv_idx
            break
    
    for row_idx, row in enumerate(data_rows, 2):
        for csv_idx, value in enumerate(row):
            if csv_idx in csv_to_col:
                col_idx = csv_to_col[csv_idx]
                ws.cell(row=row_idx, column=col_idx).value = value
        
        # Apply formatting
        is_alt_row = (row_idx % 2 == 0)
        apply_row_formatting(ws, row_idx, is_alt_row)
        
        # Track groups for merging
        url_value = row[url_col_in_csv].strip() if url_col_in_csv is not None and len(row) > url_col_in_csv else ''
        
        if url_value:
            # New group starts
            if current_group_start is not None:
                groups.append((current_group_start, row_idx - 1))
            current_group_start = row_idx
        # If url_value is empty, it's a continuation of the current group
    
    # Don't forget the last group
    if current_group_start is not None:
        last_row = len(data_rows) + 1
        groups.append((current_group_start, last_row))
    
    # Merge cells for each group with multiple rows
    for start_row, end_row in groups:
        if end_row > start_row:
            merge_cells_for_group(ws, start_row, end_row, col_indices)
    
    wb.save(excel_path)
    wb.close()
    
    print(f"✅ CSV başarıyla Excel'e dönüştürüldü: {excel_path}")
    return excel_path


def get_reported_reviews_for_business(excel_path: str, business_name: str, prefix_length: int = 25) -> set[str]:
    """Get previously reported review text prefixes for a business.
    
    Searches the Excel file for rows with matching business name that have
    been processed (have a report_id). Returns the first N characters of
    each review text for comparison.
    
    Args:
        excel_path: Path to the Excel file.
        business_name: Name of the business to search for (case-insensitive).
        prefix_length: Number of characters to use for matching (default: 25).
        
    Returns:
        Set of review text prefixes (first N characters, lowercase) for matching.
    """
    if not os.path.exists(excel_path):
        return set()
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Find column indices from header row
        col_indices = {}
        for col_idx, cell in enumerate(ws[1], 1):
            header_value = str(cell.value or '').lower().strip()
            for col_key, col_config in COLUMNS.items():
                if header_value == col_config['header'].lower() or header_value == col_key:
                    col_indices[col_key] = col_idx
                    break
        
        required_cols = ['business_name', 'report_id', 'review_text']
        if not all(col in col_indices for col in required_cols):
            wb.close()
            return set()
        
        reported_prefixes = set()
        business_name_lower = business_name.lower().strip()
        current_business_name = None
        current_has_report = False
        
        for row_idx in range(2, ws.max_row + 1):
            # Get business name (may be None if merged)
            business_cell = ws.cell(row=row_idx, column=col_indices['business_name'])
            cell_business_name = str(business_cell.value or '').strip()
            
            # If cell has a value, update current business name
            if cell_business_name:
                current_business_name = cell_business_name.lower()
                # Check if this row has report_id
                report_id_cell = ws.cell(row=row_idx, column=col_indices['report_id'])
                current_has_report = bool(report_id_cell.value)
            
            # Skip if no business name context or doesn't match
            if not current_business_name or current_business_name != business_name_lower:
                continue
            
            # Skip if the group doesn't have a report_id
            # For merged cells, we check the first row's report_id
            if not current_has_report:
                # Check current row's report_id as well (for non-merged scenarios)
                report_id_cell = ws.cell(row=row_idx, column=col_indices['report_id'])
                if not report_id_cell.value:
                    continue
            
            # Get review text
            review_text_cell = ws.cell(row=row_idx, column=col_indices['review_text'])
            review_text = str(review_text_cell.value or '').strip()
            
            if review_text:
                # Store the prefix (or full text if shorter than prefix_length)
                prefix = review_text[:prefix_length].lower()
                reported_prefixes.add(prefix)
        
        wb.close()
        return reported_prefixes
        
    except Exception as e:
        print(f"⚠️ Raporlanmış yorumlar okunurken hata: {e}")
        return set()


def get_pending_reviews(excel_path: str) -> list[dict]:
    """Get all reviews with 'beklemede' (pending) status from Excel.
    
    Args:
        excel_path: Path to the Excel file.
        
    Returns:
        List of dictionaries with review information:
        - row_idx: Row index in Excel (for updating)
        - review_url: URL of the review
        - reviewer_name: Name of the reviewer
        - review_text: Text of the review
    """
    if not os.path.exists(excel_path):
        return []
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Find column indices from header row
        col_indices = {}
        for col_idx, cell in enumerate(ws[1], 1):
            header_value = str(cell.value or '').lower().strip()
            for col_key, col_config in COLUMNS.items():
                if header_value == col_config['header'].lower() or header_value == col_key:
                    col_indices[col_key] = col_idx
                    break
        
        required_cols = ['review_url', 'reviewer_name', 'review_text', 'status']
        if not all(col in col_indices for col in required_cols):
            wb.close()
            return []
        
        pending_reviews = []
        
        for row_idx in range(2, ws.max_row + 1):
            # Get status
            status_cell = ws.cell(row=row_idx, column=col_indices['status'])
            status_value = str(status_cell.value or '').lower().strip()
            
            # Only process 'beklemede' rows
            if status_value != 'beklemede':
                continue
            
            # Get review URL
            review_url_cell = ws.cell(row=row_idx, column=col_indices['review_url'])
            review_url = str(review_url_cell.value or '').strip()
            
            if not review_url:
                continue
            
            # Get reviewer name
            reviewer_name_cell = ws.cell(row=row_idx, column=col_indices['reviewer_name'])
            reviewer_name = str(reviewer_name_cell.value or '').strip()
            
            # Get review text
            review_text_cell = ws.cell(row=row_idx, column=col_indices['review_text'])
            review_text = str(review_text_cell.value or '').strip()
            
            pending_reviews.append({
                'row_idx': row_idx,
                'review_url': review_url,
                'reviewer_name': reviewer_name,
                'review_text': review_text,
            })
        
        wb.close()
        return pending_reviews
        
    except Exception as e:
        print(f"⚠️ Beklemedeki yorumlar okunurken hata: {e}")
        return []


def update_review_status(excel_path: str, row_idx: int, new_status: str) -> bool:
    """Update the status of a review at a specific row.
    
    Args:
        excel_path: Path to the Excel file.
        row_idx: Row index to update.
        new_status: New status value (e.g., 'silindi', 'reddedildi').
        
    Returns:
        True if update was successful, False otherwise.
    """
    if not os.path.exists(excel_path):
        return False
    
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Find status column index
        status_col_idx = None
        for col_idx, cell in enumerate(ws[1], 1):
            header_value = str(cell.value or '').lower().strip()
            if header_value == COLUMNS['status']['header'].lower() or header_value == 'status':
                status_col_idx = col_idx
                break
        
        if not status_col_idx:
            wb.close()
            return False
        
        # Update the status cell
        status_cell = ws.cell(row=row_idx, column=status_col_idx)
        status_cell.value = new_status
        
        # Apply status coloring
        if new_status.lower() in STATUS_COLORS:
            status_cell.fill = STATUS_COLORS[new_status.lower()]
        
        wb.save(excel_path)
        wb.close()
        return True
        
    except Exception as e:
        print(f"⚠️ Durum güncellenirken hata: {e}")
        return False


def check_login_required(excel_path: str) -> tuple[bool, Optional[str]]:
    """Check if 'login' is written in the URL column and get the last restaurant's maps link.
    
    Scans the URL column for the word 'login' (case-insensitive). If found,
    returns the last valid maps URL from the file.
    
    Args:
        excel_path: Path to the Excel file.
        
    Returns:
        Tuple of (login_required, last_maps_url):
        - login_required: True if 'login' is found in URL column
        - last_maps_url: The last valid maps URL, or None if not found
    """
    if not os.path.exists(excel_path):
        return (False, None)
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Find URL column index from header row
        url_col_idx = None
        for col_idx, cell in enumerate(ws[1], 1):
            header_value = str(cell.value or '').lower().strip()
            if header_value == 'url' or header_value == COLUMNS['url']['header'].lower():
                url_col_idx = col_idx
                break
        
        if url_col_idx is None:
            wb.close()
            return (False, None)
        
        login_required = False
        last_maps_url = None
        
        # Scan all rows in URL column
        for row_idx in range(2, ws.max_row + 1):
            url_cell = ws.cell(row=row_idx, column=url_col_idx)
            url_value = str(url_cell.value or '').strip()
            
            if not url_value:
                continue
            
            # Check if this cell contains "login"
            if 'login' in url_value.lower():
                login_required = True
            # Check if it's a valid maps URL
            elif url_value.startswith('https://maps.') or url_value.startswith('https://www.google.com/maps') or url_value.startswith('https://goo.gl/maps'):
                last_maps_url = url_value
        
        wb.close()
        return (login_required, last_maps_url)
        
    except Exception as e:
        print(f"⚠️ Login kontrolü sırasında hata: {e}")
        return (False, None)


def refresh_formatting(excel_path: str) -> bool:
    """Refresh formatting on an existing Excel file.
    
    Applies header styling, alternating row colors, and status-based coloring.
    
    Args:
        excel_path: Path to the Excel file.
        
    Returns:
        True if successful, False otherwise.
    """
    if not os.path.exists(excel_path):
        return False
    
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Apply header formatting
        for col_idx in range(1, len(COLUMN_ORDER) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
            
            col_letter = get_column_letter(col_idx)
            if col_idx <= len(COLUMN_ORDER):
                col_key = COLUMN_ORDER[col_idx - 1]
                ws.column_dimensions[col_letter].width = COLUMNS[col_key]['width']
        
        # Apply data row formatting
        for row_idx in range(2, ws.max_row + 1):
            is_alt_row = (row_idx % 2 == 0)
            apply_row_formatting(ws, row_idx, is_alt_row)
        
        # Freeze header
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 25
        
        # Re-add status validation
        status_col_idx = COLUMN_ORDER.index('status') + 1
        status_validation = DataValidation(
            type='list',
            formula1=f'"{",".join(STATUS_OPTIONS)}"',
            allow_blank=True
        )
        status_col_letter = get_column_letter(status_col_idx)
        status_validation.add(f'{status_col_letter}2:{status_col_letter}1000')
        ws.add_data_validation(status_validation)
        
        wb.save(excel_path)
        wb.close()
        
        print(f"✅ Excel formatlaması yenilendi: {excel_path}")
        return True
        
    except Exception as e:
        print(f"⚠️ Formatlama hatası: {e}")
        return False
