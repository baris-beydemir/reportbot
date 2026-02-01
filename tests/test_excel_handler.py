"""Tests for Excel handler module."""
import os
import tempfile
import pytest
from openpyxl import load_workbook

from src.excel_handler import (
    create_formatted_workbook,
    read_excel_urls_with_count,
    update_excel_with_report,
    convert_csv_to_excel,
    refresh_formatting,
    merge_cells_for_group,
    get_reported_reviews_for_business,
    COLUMN_ORDER,
    COLUMNS,
    STATUS_OPTIONS,
    MERGE_COLUMNS,
)
from src.models import Business, Review


class TestCreateFormattedWorkbook:
    """Tests for create_formatted_workbook function."""
    
    def test_creates_workbook_with_headers(self):
        """Test that workbook is created with correct headers."""
        wb = create_formatted_workbook()
        ws = wb.active
        
        # Check all headers are present
        for col_idx, col_key in enumerate(COLUMN_ORDER, 1):
            cell = ws.cell(row=1, column=col_idx)
            assert cell.value == COLUMNS[col_key]['header']
        
        wb.close()
    
    def test_headers_have_styling(self):
        """Test that headers have proper styling applied."""
        wb = create_formatted_workbook()
        ws = wb.active
        
        # Check first header cell has styling
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.fill.start_color.rgb == '001F4E79'  # Dark blue (with alpha)
        
        wb.close()
    
    def test_freeze_panes_set(self):
        """Test that freeze panes is set on header row."""
        wb = create_formatted_workbook()
        ws = wb.active
        
        assert ws.freeze_panes == 'A2'
        
        wb.close()
    
    def test_status_validation_added(self):
        """Test that status column has data validation."""
        wb = create_formatted_workbook()
        ws = wb.active
        
        # Check that data validations exist
        assert len(ws.data_validations.dataValidation) > 0
        
        wb.close()


class TestReadExcelUrlsWithCount:
    """Tests for read_excel_urls_with_count function."""
    
    def test_reads_urls_from_excel(self):
        """Test reading URLs from Excel file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            # Create test Excel file
            wb = create_formatted_workbook()
            ws = wb.active
            
            # Add test data (count is now column 2)
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/test1'
            ws.cell(row=2, column=2).value = 2  # count column
            ws.cell(row=3, column=1).value = 'https://maps.app.goo.gl/test2'
            ws.cell(row=3, column=2).value = 3
            
            wb.save(temp_path)
            wb.close()
            
            # Read URLs
            urls = read_excel_urls_with_count(temp_path)
            
            assert len(urls) == 2
            assert urls[0] == ('https://maps.app.goo.gl/test1', 2)
            assert urls[1] == ('https://maps.app.goo.gl/test2', 3)
        finally:
            os.unlink(temp_path)
    
    def test_skips_processed_rows(self):
        """Test that rows with report_id are skipped."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            # Row with report_id (should be skipped)
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/processed'
            ws.cell(row=2, column=4).value = 'REPORT-123'  # report_id column
            
            # Row without report_id (should be included)
            ws.cell(row=3, column=1).value = 'https://maps.app.goo.gl/pending'
            
            wb.save(temp_path)
            wb.close()
            
            urls = read_excel_urls_with_count(temp_path)
            
            assert len(urls) == 1
            assert urls[0][0] == 'https://maps.app.goo.gl/pending'
        finally:
            os.unlink(temp_path)
    
    def test_default_count_is_one(self):
        """Test that count defaults to 1 if not specified."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/test'
            # No count specified
            
            wb.save(temp_path)
            wb.close()
            
            urls = read_excel_urls_with_count(temp_path)
            
            assert urls[0] == ('https://maps.app.goo.gl/test', 1)
        finally:
            os.unlink(temp_path)
    
    def test_file_not_found_raises_error(self):
        """Test that FileNotFoundError is raised for missing file."""
        with pytest.raises(FileNotFoundError):
            read_excel_urls_with_count('/nonexistent/path.xlsx')


class TestUpdateExcelWithReport:
    """Tests for update_excel_with_report function."""
    
    def test_updates_row_with_report_id(self):
        """Test updating a row with report ID."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/test'
            
            wb.save(temp_path)
            wb.close()
            
            # Update with report
            result = update_excel_with_report(
                temp_path,
                'https://maps.app.goo.gl/test',
                'REPORT-456'
            )
            
            assert result is True
            
            # Verify update
            wb = load_workbook(temp_path)
            ws = wb.active
            assert ws.cell(row=2, column=4).value == 'REPORT-456'  # report_id column
            wb.close()
        finally:
            os.unlink(temp_path)
    
    def test_updates_with_business_and_reviews(self):
        """Test updating with business and review details."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/test'
            
            wb.save(temp_path)
            wb.close()
            
            business = Business(name='Test Cafe', maps_url='https://maps.app.goo.gl/test')
            reviews = [
                Review(
                    author_name='John Doe',
                    rating=1,
                    text='Bad service',
                    review_url='https://maps.app.goo.gl/review1'
                )
            ]
            
            result = update_excel_with_report(
                temp_path,
                'https://maps.app.goo.gl/test',
                'REPORT-789',
                reviews=reviews,
                business=business
            )
            
            assert result is True
            
            # Verify all fields (column order: url=1, count=2, business_name=3, report_id=4, review_url=5, report_date=6, reviewer_name=7, ...)
            wb = load_workbook(temp_path)
            ws = wb.active
            assert ws.cell(row=2, column=3).value == 'Test Cafe'  # business_name (column 3)
            assert ws.cell(row=2, column=4).value == 'REPORT-789'  # report_id
            assert ws.cell(row=2, column=5).value == 'https://maps.app.goo.gl/review1'  # review_url
            assert ws.cell(row=2, column=6).value is not None  # report_date (should be today's date)
            assert ws.cell(row=2, column=7).value == 'John Doe'  # reviewer_name
            assert ws.cell(row=2, column=8).value == 'Bad service'  # review_text
            assert ws.cell(row=2, column=9).value == 1  # rating
            assert ws.cell(row=2, column=10).value == 'beklemede'  # status
            wb.close()
        finally:
            os.unlink(temp_path)
    
    def test_returns_false_for_nonexistent_url(self):
        """Test that False is returned when URL not found."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/existing'
            
            wb.save(temp_path)
            wb.close()
            
            result = update_excel_with_report(
                temp_path,
                'https://maps.app.goo.gl/nonexistent',
                'REPORT-000'
            )
            
            assert result is False
        finally:
            os.unlink(temp_path)
    
    def test_merges_cells_for_multiple_reviews(self):
        """Test that cells are merged when multiple reviews are added."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/test'
            
            wb.save(temp_path)
            wb.close()
            
            business = Business(name='Test Cafe', maps_url='https://maps.app.goo.gl/test')
            reviews = [
                Review(author_name='User1', rating=1, text='Bad', review_url='https://review1'),
                Review(author_name='User2', rating=2, text='Meh', review_url='https://review2'),
            ]
            
            result = update_excel_with_report(
                temp_path,
                'https://maps.app.goo.gl/test',
                'REPORT-MERGE',
                reviews=reviews,
                business=business
            )
            
            assert result is True
            
            # Verify cells are merged
            wb = load_workbook(temp_path)
            ws = wb.active
            
            merged_ranges = [str(r) for r in ws.merged_cells.ranges]
            
            # URL, count, business_name, report_id should be merged (A2:A3, B2:B3, C2:C3, D2:D3)
            assert 'A2:A3' in merged_ranges
            assert 'B2:B3' in merged_ranges
            assert 'D2:D3' in merged_ranges  # report_id
            
            # Verify both reviews are in the sheet (review_url is column 5)
            assert ws.cell(row=2, column=5).value == 'https://review1'  # First review
            assert ws.cell(row=3, column=5).value == 'https://review2'  # Second review
            
            wb.close()
        finally:
            os.unlink(temp_path)


class TestConvertCsvToExcel:
    """Tests for convert_csv_to_excel function."""
    
    def test_converts_csv_with_header(self):
        """Test converting CSV file with header row."""
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False, mode='w', encoding='utf-8') as f:
            f.write('url;count;status\n')
            f.write('https://maps.app.goo.gl/test1;2;beklemede\n')
            f.write('https://maps.app.goo.gl/test2;1;silindi\n')
            csv_path = f.name
        
        excel_path = csv_path.replace('.csv', '.xlsx')
        
        try:
            result = convert_csv_to_excel(csv_path)
            
            assert result == excel_path
            assert os.path.exists(excel_path)
            
            # Verify content
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Check header
            assert ws.cell(row=1, column=1).value == 'URL'
            
            # Check data (count is now column 2)
            assert ws.cell(row=2, column=1).value == 'https://maps.app.goo.gl/test1'
            assert ws.cell(row=2, column=2).value == '2'  # count column
            
            wb.close()
        finally:
            os.unlink(csv_path)
            if os.path.exists(excel_path):
                os.unlink(excel_path)
    
    def test_merges_cells_for_multiple_reviews(self):
        """Test that cells are merged when multiple reviews exist for same restaurant."""
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False, mode='w', encoding='utf-8') as f:
            f.write('url;business_name;count;report_id;review_url;reviewer_name;review_text;rating;status\n')
            f.write('https://maps.app.goo.gl/test1;Test Cafe;2;REP-123;https://review1;User1;Good;5;beklemede\n')
            f.write(';;;https://review2;User2;Bad;1;beklemede\n')  # Second review, same restaurant
            f.write('https://maps.app.goo.gl/test2;Other Place;1;REP-456;https://review3;User3;OK;3;silindi\n')
            csv_path = f.name
        
        excel_path = csv_path.replace('.csv', '.xlsx')
        
        try:
            convert_csv_to_excel(csv_path)
            
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Check that cells are merged for the first restaurant (rows 2-3)
            merged_ranges = [str(r) for r in ws.merged_cells.ranges]
            
            # URL, business_name, count, report_id should be merged
            assert 'A2:A3' in merged_ranges
            assert 'B2:B3' in merged_ranges
            assert 'C2:C3' in merged_ranges
            assert 'D2:D3' in merged_ranges
            
            # Row 4 (second restaurant) should not have merged cells
            assert 'A4:A5' not in merged_ranges
            
            wb.close()
        finally:
            os.unlink(csv_path)
            if os.path.exists(excel_path):
                os.unlink(excel_path)
    
    def test_file_not_found_raises_error(self):
        """Test that FileNotFoundError is raised for missing CSV."""
        with pytest.raises(FileNotFoundError):
            convert_csv_to_excel('/nonexistent/path.csv')


class TestRefreshFormatting:
    """Tests for refresh_formatting function."""
    
    def test_refreshes_header_styling(self):
        """Test that header styling is refreshed."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            # Create basic workbook without styling
            wb = create_formatted_workbook()
            ws = wb.active
            
            # Add some data
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/test'
            ws.cell(row=2, column=9).value = 'beklemede'
            
            # Remove styling from header (simulate)
            ws.cell(row=1, column=1).font = None
            
            wb.save(temp_path)
            wb.close()
            
            # Refresh formatting
            result = refresh_formatting(temp_path)
            
            assert result is True
            
            # Verify styling is back
            wb = load_workbook(temp_path)
            ws = wb.active
            assert ws.cell(row=1, column=1).font.bold is True
            wb.close()
        finally:
            os.unlink(temp_path)
    
    def test_returns_false_for_nonexistent_file(self):
        """Test that False is returned for missing file."""
        result = refresh_formatting('/nonexistent/path.xlsx')
        assert result is False


class TestGetReportedReviewsForBusiness:
    """Tests for get_reported_reviews_for_business function."""
    
    def test_returns_reported_reviews_for_matching_business(self):
        """Test finding previously reported reviews for a business name."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            # Add a reported review for "Test Cafe"
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/old1'
            ws.cell(row=2, column=3).value = 'Test Cafe'  # business_name
            ws.cell(row=2, column=4).value = 'REPORT-OLD'  # report_id
            ws.cell(row=2, column=8).value = 'This review has bad content and should be ignored'  # review_text
            
            # Add another reported review for same business
            ws.cell(row=3, column=1).value = ''
            ws.cell(row=3, column=4).value = ''  # Same report
            ws.cell(row=3, column=8).value = 'Another bad review with different text'  # review_text
            
            # Add a review for different business (should not be returned)
            ws.cell(row=4, column=1).value = 'https://maps.app.goo.gl/other'
            ws.cell(row=4, column=3).value = 'Other Restaurant'
            ws.cell(row=4, column=4).value = 'REPORT-OTHER'
            ws.cell(row=4, column=8).value = 'Some review for other place'
            
            wb.save(temp_path)
            wb.close()
            
            # Get reported reviews for "Test Cafe"
            reported = get_reported_reviews_for_business(temp_path, 'Test Cafe')
            
            assert len(reported) == 2
            assert 'this review has bad conte' in reported  # First 25 chars, lowercase
            assert 'another bad review with d' in reported  # First 25 chars, lowercase
        finally:
            os.unlink(temp_path)
    
    def test_returns_empty_for_no_matches(self):
        """Test that empty set is returned when no matching business found."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/test'
            ws.cell(row=2, column=3).value = 'Other Business'
            ws.cell(row=2, column=4).value = 'REPORT-123'
            ws.cell(row=2, column=8).value = 'Some review text here'
            
            wb.save(temp_path)
            wb.close()
            
            reported = get_reported_reviews_for_business(temp_path, 'Test Cafe')
            
            assert len(reported) == 0
        finally:
            os.unlink(temp_path)
    
    def test_case_insensitive_business_name_matching(self):
        """Test that business name matching is case-insensitive."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/test'
            ws.cell(row=2, column=3).value = 'TEST CAFE'  # Uppercase
            ws.cell(row=2, column=4).value = 'REPORT-123'
            ws.cell(row=2, column=8).value = 'Review text for case test'
            
            wb.save(temp_path)
            wb.close()
            
            # Search with different case
            reported = get_reported_reviews_for_business(temp_path, 'test cafe')
            
            assert len(reported) == 1
            assert 'review text for case test' in reported  # lowercase, full text since < 25 chars
        finally:
            os.unlink(temp_path)
    
    def test_skips_rows_without_report_id(self):
        """Test that rows without report_id are skipped."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            # Row with report_id (should be included)
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/test1'
            ws.cell(row=2, column=3).value = 'Test Cafe'
            ws.cell(row=2, column=4).value = 'REPORT-123'
            ws.cell(row=2, column=8).value = 'Reported review text here'
            
            # Row without report_id (should be skipped)
            ws.cell(row=3, column=1).value = 'https://maps.app.goo.gl/test2'
            ws.cell(row=3, column=3).value = 'Test Cafe'
            ws.cell(row=3, column=4).value = ''  # No report_id
            ws.cell(row=3, column=8).value = 'Not reported yet'
            
            wb.save(temp_path)
            wb.close()
            
            reported = get_reported_reviews_for_business(temp_path, 'Test Cafe')
            
            assert len(reported) == 1
            assert 'reported review text here' in reported  # lowercase
        finally:
            os.unlink(temp_path)
    
    def test_returns_empty_for_nonexistent_file(self):
        """Test that empty set is returned for missing file."""
        reported = get_reported_reviews_for_business('/nonexistent/path.xlsx', 'Test')
        assert len(reported) == 0
    
    def test_handles_short_review_text(self):
        """Test handling of review text shorter than 25 characters."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/test'
            ws.cell(row=2, column=3).value = 'Test Cafe'
            ws.cell(row=2, column=4).value = 'REPORT-123'
            ws.cell(row=2, column=8).value = 'Short text'  # Less than 25 chars
            
            wb.save(temp_path)
            wb.close()
            
            reported = get_reported_reviews_for_business(temp_path, 'Test Cafe')
            
            assert len(reported) == 1
            assert 'short text' in reported  # lowercase
        finally:
            os.unlink(temp_path)
    
    def test_handles_merged_cells_business_name(self):
        """Test that business name is correctly read from merged cells."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        
        try:
            wb = create_formatted_workbook()
            ws = wb.active
            
            # First row has business name, subsequent rows are merged
            ws.cell(row=2, column=1).value = 'https://maps.app.goo.gl/test'
            ws.cell(row=2, column=3).value = 'Test Cafe'
            ws.cell(row=2, column=4).value = 'REPORT-123'
            ws.cell(row=2, column=8).value = 'First review text here'
            
            # Second row (merged business_name cell)
            ws.cell(row=3, column=3).value = None  # Merged cell - None
            ws.cell(row=3, column=4).value = None  # Same report
            ws.cell(row=3, column=8).value = 'Second review different'
            
            # Merge the business_name cells
            ws.merge_cells('C2:C3')
            ws.merge_cells('D2:D3')
            
            wb.save(temp_path)
            wb.close()
            
            reported = get_reported_reviews_for_business(temp_path, 'Test Cafe')
            
            # Both reviews should be found
            assert len(reported) == 2
            assert 'first review text here' in reported  # lowercase
            assert 'second review different' in reported  # lowercase
        finally:
            os.unlink(temp_path)
